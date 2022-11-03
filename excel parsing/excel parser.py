from openpyxl import load_workbook

wb = load_workbook(filename= 'Пример для конкурса_пул_03.11.2022_.xlsx', data_only= True)

ws = wb.active

list_info = []
for i in range(2, ws.max_row + 1):
    list_info.append([
    ws['A' + str(i)].value, ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value,
    ws['D' + str(i)].value, ws['E' + str(i)].value, ws['F' + str(i)].value, ws['G' + str(i)].value, 
    ws['H' + str(i)].value, ws['I' + str(i)].value, ws['J' + str(i)].value, ws['K' + str(i)].value])

for i in range(len(list_info)):
    print(list_info[i])